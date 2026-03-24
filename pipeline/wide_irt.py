"""
pipeline/wide_irt.py
Procesa base bruta IRT → formato Wide (_IRT1/_IRT2/_IRT3).
Misma interfaz que wide_top.procesar_wide.
"""
import pandas as pd
import numpy as np
import unicodedata, re, warnings
from io import BytesIO
warnings.filterwarnings('ignore')

def _norm_str(s):
    return unicodedata.normalize('NFD', str(s).lower()).encode('ascii','ignore').decode()

MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
            7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

def auto_col(cols, keywords, nombre_col):
    for c in cols:
        if any(_norm_str(k) in _norm_str(c) for k in keywords):
            return c
    raise ValueError(f"Columna '{nombre_col}' no encontrada.")

def _parse_fecha(serie):
    if pd.api.types.is_datetime64_any_dtype(serie): return serie
    _MES = {'ene':'Jan','feb':'Feb','mar':'Mar','abr':'Apr','may':'May','jun':'Jun',
            'jul':'Jul','ago':'Aug','sept':'Sep','sep':'Sep','oct':'Oct','nov':'Nov','dic':'Dec'}
    def _conv(val):
        s = str(val).strip().lower()
        for es,en in _MES.items(): s = re.sub(rf'\b{es}\b', en, s)
        return pd.to_datetime(s, errors='coerce')
    result = pd.to_datetime(serie, errors='coerce')
    mask = result.isna() & serie.notna()
    if mask.any(): result[mask] = serie[mask].apply(_conv)
    return result

def norm_sust_v3(s):
    if pd.isna(s): return None
    raw = str(s).strip()
    if raw in ('0',''): return None
    raw = re.split(r'[\r\n]', raw)[0].strip()
    raw = re.sub(r'\(.*?\)','',raw).strip()
    raw = re.sub(r'^(las dos|ambas|los dos|ambos)[,\s]+','',raw,flags=re.IGNORECASE).strip()
    primera = re.split(r'\s+y\s+|[/,+]',raw,maxsplit=1)[0].strip()
    n = _norm_str(primera)
    if any(x in n for x in ['ninguno','ninguna','niega','no aplica','no consume','nada']): return None
    if any(x in n for x in ['ludopatia','juego','apuesta']): return None
    if any(x in n for x in ['alcohol','alchol','cerveza','licor','aguard','ron']): return 'Alcohol'
    if any(x in n for x in ['marihu','marhuana','cannabis','marij','weed','crispy']): return 'Marihuana'
    if any(x in n for x in ['tusi','tussi','tusy','tuci','2cb']): return 'Tusi'
    if any(x in n for x in ['pasta base','papelillo','pbc','basuco','bazuco']): return 'Pasta Base/Basuco'
    if any(x in n for x in ['metanfet','anfetam','cristal','crystal']): return 'Metanfetamina'
    if any(x in n for x in ['crack','piedra','paco']): return 'Crack'
    if any(x in n for x in ['cocain','cocai','perico','coke']): return 'Cocaína'
    if any(x in n for x in ['tabaco','cigarr','nicot']): return 'Tabaco'
    if any(x in n for x in ['inhalant','thiner','activo','pegamento']): return 'Inhalantes'
    if any(x in n for x in ['sedant','benzod','tranqui','clonaz','rivotril']): return 'Sedantes'
    if any(x in n for x in ['opiod','heroina','morfin','fentanil','tramad']): return 'Opiáceos'
    if any(x in n for x in ['extasis','mdma']): return 'Éxtasis'
    if any(x in n for x in ['ketam']): return 'Ketamina'
    return None

_SUST_KEYS = [_norm_str(x) for x in
    ['sustancia principal','cual considera','cuál considera','genera mas problemas']]


def procesar_wide(input_path, filtro_centro=None, fecha_desde=None, fecha_hasta=None):
    logs = []
    df = pd.read_excel(input_path, sheet_name=0, header=0)
    logs.append(f"✓ {len(df)} filas × {len(df.columns)} columnas")

    COL_CODIGO = auto_col(df.columns,
        ['identificacion','identificación','2 primeras letras','primer nombre','cod_pac'],
        'Código paciente')
    COL_FECHA = auto_col(df.columns,
        ['fecha de administracion','fecha_administracion','fecha administracion',
         'fecha entrevista','fecha_entrevista'],
        'Fecha de Administración')

    hoy = pd.Timestamp.now()
    df[COL_FECHA] = _parse_fecha(df[COL_FECHA])
    alertas = []

    COL_CENTRO = None
    for c in df.columns:
        nc = _norm_str(c)
        if any(k in nc for k in ['codigo del centro','centro de tratamiento','servicio de tratamiento']):
            if 'trabajo' not in nc and 'estudio' not in nc:
                COL_CENTRO = c; break

    if filtro_centro and COL_CENTRO:
        n_a = len(df)
        df = df[df[COL_CENTRO].astype(str).str.strip()==filtro_centro.strip()].copy().reset_index(drop=True)
        logs.append(f"✓ Filtro centro: {n_a}→{len(df)}")
        if len(df)==0: raise ValueError(f"Centro '{filtro_centro}' sin registros.")

    if fecha_desde or fecha_hasta:
        mask = pd.Series([True]*len(df), index=df.index)
        if fecha_desde: mask &= df[COL_FECHA] >= pd.Timestamp(fecha_desde+'-01')
        if fecha_hasta: mask &= df[COL_FECHA] <= pd.Timestamp(fecha_hasta+'-01') + pd.offsets.MonthEnd(0)
        n_a = len(df); df = df[mask].copy().reset_index(drop=True)
        logs.append(f"✓ Filtro período: {n_a}→{len(df)}")
        if len(df)==0: raise ValueError("Sin registros en el período.")

    centro_lookup = df.groupby(COL_CODIGO)[COL_CENTRO].first().to_dict() if COL_CENTRO else {}
    def get_centro(cod): return str(centro_lookup.get(cod,'—'))[:60]

    COL_FN = next((c for c in df.columns if c != COL_CODIGO and
                   any(k in _norm_str(c) for k in ['fecha de nacimiento','fecha_nacimiento'])), None)
    if COL_FN:
        df[COL_FN] = _parse_fecha(df[COL_FN])
        for idx,row in df.iterrows():
            fn=row[COL_FN]; cod=row[COL_CODIGO]
            if pd.isna(fn): continue
            if fn>hoy:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,'Valor':str(fn.date()),'Regla':'Fecha nacimiento futura'})
                df.at[idx,COL_FN]=np.nan; continue
            edad=(hoy-fn).days/365.25
            if edad<10 or edad>100:
                alertas.append({'Código':cod,'Centro':get_centro(cod),'Columna':COL_FN,'Valor':str(fn.date()),'Regla':f'Edad={edad:.1f} años'})
                df.at[idx,COL_FN]=np.nan

    for c in [c for c in df.columns if '(0-7)' in c and 'Promedio' not in c]:
        num=pd.to_numeric(df[c],errors='coerce'); mask_=num>7
        for idx in df[mask_].index:
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),'Columna':c,'Valor':df.at[idx,c],'Regla':'Días sem>7'})
            df.at[idx,c]=np.nan

    for c in [c for c in df.columns if 'Total (0-28)' in c and 'Promedio' not in c]:
        num=pd.to_numeric(df[c],errors='coerce')
        for idx in df[(num>28)|(num<0)].index:
            alertas.append({'Código':df.at[idx,COL_CODIGO],'Centro':get_centro(df.at[idx,COL_CODIGO]),'Columna':c,'Valor':df.at[idx,c],'Regla':'Días mes fuera 0-28'})
            df.at[idx,c]=np.nan

    logs.append(f"✓ {len(alertas)} valores corregidos")

    fechas_ok = df[COL_FECHA].dropna()
    fechas_ok = fechas_ok[(fechas_ok.dt.year>=hoy.year-10)&(fechas_ok.dt.year<=hoy.year+1)]
    if len(fechas_ok):
        f0,f1=fechas_ok.min(),fechas_ok.max()
        if f0.year==f1.year and f0.month==f1.month: periodo=f'{MESES_ES[f0.month]} {f0.year}'
        elif f0.year==f1.year: periodo=f'{MESES_ES[f0.month]}–{MESES_ES[f1.month]} {f0.year}'
        else: periodo=f'{MESES_ES[f0.month]} {f0.year} – {MESES_ES[f1.month]} {f1.year}'
    else: periodo='Período no determinado'

    df = df.sort_values([COL_CODIGO,COL_FECHA]).reset_index(drop=True)
    conteo  = df[COL_CODIGO].value_counts()
    N_total = int(conteo.shape[0])
    N_irt2  = int((conteo>=2).sum())
    N_irt3  = int((conteo>=3).sum())
    N_solo1 = N_total - N_irt2

    rows1,rows2,rows3=[],[],[]
    for cod,grp in df.groupby(COL_CODIGO,sort=False):
        grp=grp.reset_index(drop=True)
        rows1.append(grp.loc[0])
        if len(grp)>=2: rows2.append(grp.loc[1])
        if len(grp)>=3: rows3.append(grp.loc[2])

    df1=pd.DataFrame(rows1).reset_index(drop=True)
    otras=[c for c in df1.columns if c!=COL_CODIGO]
    t1=df1.rename(columns={c:f'{c}_IRT1' for c in otras})
    wide=t1.copy()

    if rows2:
        df2=pd.DataFrame(rows2).reset_index(drop=True)
        df2a=df2.set_index(COL_CODIGO).reindex(df1[COL_CODIGO]).reset_index()
        t2=df2a.rename(columns={c:f'{c}_IRT2' for c in otras})
        wide=wide.merge(t2,on=COL_CODIGO,how='left')
    if rows3:
        df3=pd.DataFrame(rows3).reset_index(drop=True)
        df3a=df3.set_index(COL_CODIGO).reindex(df1[COL_CODIGO]).reset_index()
        t3=df3a.rename(columns={c:f'{c}_IRT3' for c in otras})
        wide=wide.merge(t3,on=COL_CODIGO,how='left')

    wide.insert(1,'Tiene_IRT1','Sí')
    irt2_cols=[c for c in wide.columns if c.endswith('_IRT2')]
    wide.insert(2,'Tiene_IRT2',
        wide[irt2_cols].notna().any(axis=1).map({True:'Sí',False:'No'}) if irt2_cols else 'No')

    _col_f1=next((c for c in wide.columns if 'fecha' in _norm_str(c) and c.endswith('_IRT1')),None)
    _HOY=pd.Timestamp.now().normalize()
    _n_rojo=_n_naranja=_n_verde=0
    if _col_f1:
        _f=pd.to_datetime(wide[_col_f1],errors='coerce')
        _d=(_HOY-_f).dt.days
        def _al(d):
            if pd.isna(d): return ''
            if d<60: return '🟢 <60 dias'
            if d<90: return '🟠 60-89 dias'
            return '🔴 90+ dias'
        wide['Dias_desde_IRT1']=_d.where(wide['Tiene_IRT2']=='No',other=None)
        wide['Alerta_IRT2']=_d.where(wide['Tiene_IRT2']=='No').apply(lambda d:_al(d) if not pd.isna(d) else '')
        wide.loc[wide['Tiene_IRT2']=='Sí','Alerta_IRT2']='Completado'
        _n_rojo=int((wide['Alerta_IRT2']=='🔴 90+ dias').sum())
        _n_naranja=int((wide['Alerta_IRT2']=='🟠 60-89 dias').sum())
        _n_verde=int((wide['Alerta_IRT2']=='🟢 <60 dias').sum())
    else:
        wide['Dias_desde_IRT1']=None; wide['Alerta_IRT2']=''

    for _sfx in ('_IRT1','_IRT2','_IRT3'):
        _col=next((c for c in wide.columns if c.endswith(_sfx) and
                   any(k in _norm_str(c) for k in _SUST_KEYS) and 'RAW' not in c),None)
        if not _col: continue
        _raw=_col.replace(_sfx,f'_RAW{_sfx}')
        wide.rename(columns={_col:_raw},inplace=True)
        wide[_col]=wide[_raw].apply(norm_sust_v3)
        _i=wide.columns.get_loc(_raw)
        wide=wide[[*wide.columns[:_i+1],_col,*[c for c in wide.columns[_i+1:] if c!=_col]]]
        logs.append(f"✓ Sustancia normalizada {_sfx}")

    dupes_data=[]
    for _,row in df[df.duplicated([COL_CODIGO,COL_FECHA],keep=False)][[COL_CODIGO,COL_FECHA]].drop_duplicates().iterrows():
        dupes_data.append({'Código':row[COL_CODIGO],'Fecha':str(row[COL_FECHA])[:10]})

    excel_bytes=_excel_wide(wide,alertas,dupes_data,COL_CODIGO,COL_CENTRO,_col_f1,
                            N_total,N_irt2,N_irt3,N_solo1,len(alertas),len(dupes_data),
                            _n_rojo,_n_naranja,_n_verde,periodo)

    col_sust=next((c for c in wide.columns if any(k in _norm_str(c) for k in _SUST_KEYS)
                   and c.endswith('_IRT1') and 'RAW' not in c),None)
    sust_dist=wide[col_sust].dropna().value_counts().head(8).to_dict() if col_sust else {}

    centros=[]
    if COL_CENTRO:
        ccw=f'{COL_CENTRO}_IRT1'
        if ccw in wide.columns:
            apps=df.groupby(COL_CENTRO).size().reset_index(name='Aplicaciones').rename(columns={COL_CENTRO:'Centro'})
            res=wide.groupby(ccw).agg(Pacientes=(COL_CODIGO,'count'),Con_IRT2=('Tiene_IRT2',lambda x:(x=='Sí').sum())).reset_index().rename(columns={ccw:'Centro'})
            res['Sin_IRT2']=res['Pacientes']-res['Con_IRT2']
            if alertas:
                df_al=pd.DataFrame(alertas)
                corr=df_al.groupby('Centro').size().reset_index(name='Vals_corregidos')
                res=res.merge(corr,on='Centro',how='left')
            else: res['Vals_corregidos']=0
            res['Vals_corregidos']=res['Vals_corregidos'].fillna(0).astype(int)
            res=res.merge(apps,on='Centro',how='left')
            res['Aplicaciones']=res['Aplicaciones'].fillna(0).astype(int)
            res=res.sort_values('Aplicaciones',ascending=False)
            tots={'Centro':'TOTAL','Aplicaciones':int(res['Aplicaciones'].sum()),
                  'Pacientes':int(res['Pacientes'].sum()),'Con_IRT2':int(res['Con_IRT2'].sum()),
                  'Sin_IRT2':int(res['Sin_IRT2'].sum()),'Vals_corregidos':int(res['Vals_corregidos'].sum())}
            centros=res[['Centro','Aplicaciones','Pacientes','Con_IRT2','Sin_IRT2','Vals_corregidos']].to_dict('records')
            centros.append(tots)

    return {'wide':wide,'filtro_centro':filtro_centro,'fecha_desde':fecha_desde,'fecha_hasta':fecha_hasta,
            'stats':{'N_total':N_total,'N_irt2':N_irt2,'N_irt3':N_irt3,'N_solo1':N_solo1,
                     'pct_irt2':round(N_irt2/N_total*100,1) if N_total else 0,
                     'N_alertas':len(alertas),'N_dupes':len(dupes_data),
                     'n_rojo':_n_rojo,'n_naranja':_n_naranja,'n_verde':_n_verde,
                     'cols_wide':len(wide.columns),'sust_dist':sust_dist},
            'centros':centros,'alertas':alertas,'dupes':dupes_data,
            'periodo':periodo,'excel_bytes':excel_bytes,'logs':logs}


def _excel_wide(wide,alertas,dupes,COL_CODIGO,COL_CENTRO,col_f1,
                N_total,N_irt2,N_irt3,N_solo1,N_al,N_du,
                n_rojo,n_naranja,n_verde,periodo):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
    from openpyxl.utils import get_column_letter
    C_DARK='1F3864'; C_MID='2E75B6'; C_WHITE='FFFFFF'
    C_ALT='EEF4FB'; C_BDR='B8CCE4'; C_IRT2='00B0F0'

    wb=Workbook(); ws=wb.active; ws.title='Base Wide'
    ws.sheet_properties.tabColor=C_DARK
    ws.sheet_view.showGridLines=False; ws.freeze_panes='B3'
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(wide.columns))
    ct=ws.cell(1,1); ct.value=f'QALAT · Base Wide IRT · {periodo} · {N_total} pacientes'
    ct.font=Font(bold=True,size=10,color=C_WHITE,name='Arial')
    ct.fill=PatternFill('solid',start_color=C_DARK)
    ct.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=22
    for ci,col in enumerate(wide.columns,1):
        c=ws.cell(2,ci); c.value=col
        c.font=Font(bold=True,size=8,color=C_WHITE,name='Arial')
        c.fill=PatternFill('solid',start_color=C_DARK)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.column_dimensions[c.column_letter].width=max(10,min(35,len(str(col))*0.85))
    ws.row_dimensions[2].height=40
    for ri,row in wide.iterrows():
        bg=PatternFill('solid',start_color=(C_ALT if ri%2==0 else C_WHITE))
        for ci,val in enumerate(row,1):
            c=ws.cell(ri+3,ci)
            c.value=None if (not isinstance(val,str) and pd.isna(val)) else val
            c.font=Font(size=8,name='Arial'); c.fill=bg
            c.alignment=Alignment(horizontal='center',vertical='center')

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    wr=wb.create_sheet('Resumen')
    wr.sheet_properties.tabColor=C_MID
    wr.sheet_view.showGridLines=False
    wr.column_dimensions['A'].width=2
    wr.column_dimensions['B'].width=40
    wr.column_dimensions['C'].width=20

    def _hdr(ws,row,text,bg=C_DARK):
        ws.merge_cells(f'B{row}:C{row}')
        c=ws[f'B{row}']; c.value=text
        c.font=Font(bold=True,size=10,color=C_WHITE,name='Arial')
        c.fill=PatternFill('solid',start_color=bg)
        c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
        ws.row_dimensions[row].height=20

    def _row(ws,row,label,value,alt=False,color_val=None):
        bg='EEF4FB' if alt else C_WHITE
        ws.row_dimensions[row].height=16
        for ci,val in zip(['B','C'],[label,value]):
            c=ws[f'{ci}{row}']; c.value=val
            c.font=Font(size=9,name='Arial',bold=(ci=='C'),
                        color=color_val if (color_val and ci=='C') else '000000')
            c.fill=PatternFill('solid',start_color=bg)
            c.alignment=Alignment(horizontal='left' if ci=='B' else 'center',
                                  vertical='center',indent=1 if ci=='B' else 0)
            c.border=Border(bottom=Side(style='thin',color=C_BDR))

    wr.merge_cells('B1:C1')
    c=wr['B1']; c.value='RESUMEN  ·  Instrumento IRT'
    c.font=Font(bold=True,size=13,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color=C_DARK)
    c.alignment=Alignment(horizontal='center',vertical='center')
    wr.row_dimensions[1].height=32

    R=3
    _hdr(wr,R,'ESTADÍSTICAS DE LA BASE'); R+=1
    _row(wr,R,'Período detectado',periodo,True); R+=1
    _row(wr,R,'Total pacientes únicos (IRT1)',N_total); R+=1
    _row(wr,R,'  → Solo IRT1 (sin seguimiento)',f'{N_solo1} ({round(N_solo1/N_total*100,1) if N_total else 0}%)',True); R+=1
    _row(wr,R,'  → Con IRT2 (seguimiento 3m)',f'{N_irt2} ({round(N_irt2/N_total*100,1) if N_total else 0}%)'); R+=1
    _row(wr,R,'  → Con IRT3 (seguimiento 6m)',f'{N_irt3} ({round(N_irt3/N_total*100,1) if N_total else 0}%)',True); R+=1
    _row(wr,R,'Columnas base wide',len(wide.columns)); R+=1
    _row(wr,R,'Valores corregidos (→ NaN)',
         f'{N_al}  {"⚠ Ver hoja Alertas" if N_al else "✅ Sin errores"}',True,
         color_val='C00000' if N_al else '1A6632'); R+=2
    _hdr(wr,R,'SEMÁFORO DE SEGUIMIENTO',C_MID); R+=1
    _row(wr,R,'🔴 Urgentes (90+ días sin IRT2)',n_rojo,True,color_val='C00000' if n_rojo else None); R+=1
    _row(wr,R,'🟠 Próximos (60–89 días)',n_naranja,False,color_val='E67E22' if n_naranja else None); R+=1
    _row(wr,R,'🟢 Con tiempo (<60 días)',n_verde,True,color_val='1A6632' if n_verde else None); R+=1
    _row(wr,R,'✅ Completados (tienen IRT2)',N_irt2); R+=1

    # ── Hoja 3: Alertas ───────────────────────────────────────────────────────
    wa=wb.create_sheet('Alertas')
    wa.sheet_properties.tabColor='C00000' if alertas else '70AD47'
    wa.sheet_view.showGridLines=False
    wa.column_dimensions['A'].width=2
    wa.column_dimensions['B'].width=24
    wa.column_dimensions['C'].width=40
    wa.column_dimensions['D'].width=55
    wa.column_dimensions['E'].width=20
    wa.column_dimensions['F'].width=48
    wa.row_dimensions[1].height=28
    wa.merge_cells('B1:F1')
    c=wa['B1']
    c.value=(f'⚠  ALERTAS DE VALIDACIÓN  ·  {len(alertas)} valor(es) corregido(s) → NaN'
             if alertas else '✅  SIN ERRORES DE VALIDACIÓN')
    c.font=Font(bold=True,size=12,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color='C00000' if alertas else '70AD47')
    c.alignment=Alignment(horizontal='center',vertical='center')
    if alertas:
        wa.row_dimensions[2].height=14
        wa.merge_cells('B2:F2')
        c=wa['B2']
        c.value='El valor original fue reemplazado por vacío (NaN). El paciente sigue en la base con el resto de sus datos.'
        c.font=Font(italic=True,size=8,color='7F0000',name='Arial')
        c.fill=PatternFill('solid',start_color='FFE6E6')
        c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
        wa.row_dimensions[4].height=20
        for ci,hdr in enumerate(['Código Paciente','Centro / Servicio',
                                   'Variable / Columna','Valor Original','Regla Violada'],2):
            c=wa.cell(4,ci,hdr)
            c.font=Font(bold=True,size=9,color=C_WHITE,name='Arial')
            c.fill=PatternFill('solid',start_color='C00000')
            c.alignment=Alignment(horizontal='center',vertical='center')
            c.border=Border(bottom=Side(style='medium',color='888888'))
        for ri,alerta in enumerate(alertas,5):
            wa.row_dimensions[ri].height=15
            bg='FFF2F2' if ri%2==0 else 'FFFFFF'
            vals=[alerta.get('Código',''),alerta.get('Centro',''),
                  alerta.get('Columna',''),alerta.get('Valor',''),alerta.get('Regla','')]
            for ci,v in enumerate(vals,2):
                c=wa.cell(ri,ci,str(v))
                c.font=Font(size=8,name='Arial',bold=(ci==4),
                            color='1F3864' if ci==4 else '000000')
                c.fill=PatternFill('solid',start_color=bg)
                c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
                c.border=Border(bottom=Side(style='thin',color='DDDDDD'))

    # ── Hoja 4: Calidad de Datos ──────────────────────────────────────────────
    N_du_pac=len(set(d.get('Código','') for d in dupes)) if dupes else 0
    if N_du_pac==0: sem_col='70AD47'; sem_txt='✅  Sin problemas de fechas duplicadas'
    elif N_du_pac<=3: sem_col='FFD966'; sem_txt=f'⚠️  Atención: {N_du_pac} paciente(s) con fecha duplicada'
    else: sem_col='C00000'; sem_txt=f'🔴  Problema: {N_du_pac} pacientes con fecha duplicada'

    wq=wb.create_sheet('Calidad de Datos')
    wq.sheet_properties.tabColor=sem_col
    wq.sheet_view.showGridLines=False
    wq.column_dimensions['A'].width=2; wq.column_dimensions['B'].width=30
    wq.column_dimensions['C'].width=20; wq.column_dimensions['D'].width=16
    wq.column_dimensions['E'].width=40
    wq.row_dimensions[1].height=32
    wq.merge_cells('B1:E1')
    c=wq['B1']; c.value='🔍  CALIDAD DE DATOS  ·  Monitor de Errores de Ingreso'
    c.font=Font(bold=True,size=13,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color=C_DARK)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    wq.row_dimensions[2].height=6
    wq.row_dimensions[3].height=26
    wq.merge_cells('B3:E3')
    c=wq['B3']; c.value=sem_txt
    c.font=Font(bold=True,size=12,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color=sem_col)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=2)
    wq.row_dimensions[5].height=20
    wq.merge_cells('B5:E5')
    c=wq['B5']; c.value='  INDICADORES GENERALES'
    c.font=Font(bold=True,size=10,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color=C_DARK)
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    Rq=6
    for lbl,val,alt,cv in [
        ('Pacientes únicos',N_total,True,None),
        ('Pacientes con fecha duplicada',
         f'{N_du_pac} ({round(N_du_pac/N_total*100,1) if N_total else 0}%)',
         False,'C00000' if N_du_pac>0 else '1A6632'),
        ('Valores inválidos corregidos (→ NaN)',N_al,True,
         'C00000' if N_al else '1A6632'),
    ]:
        _row(wq,Rq,lbl,val,alt,cv); Rq+=1
    Rq+=1
    wq.row_dimensions[Rq].height=20
    wq.merge_cells(f'B{Rq}:E{Rq}')
    c=wq[f'B{Rq}']
    c.value=(f'  CASOS CON FECHA DUPLICADA  ·  {N_du_pac} paciente(s)'
             if N_du_pac>0 else '  CASOS CON FECHA DUPLICADA  ·  Ninguno ✅')
    c.font=Font(bold=True,size=10,color=C_WHITE,name='Arial')
    c.fill=PatternFill('solid',start_color='C00000' if N_du_pac>0 else '70AD47')
    c.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    Rq+=1
    if dupes:
        for ci,h in enumerate(['Código Paciente','Fecha Duplicada','N° aplicaciones'],2):
            c=wq.cell(Rq,ci,h)
            c.font=Font(bold=True,size=9,color=C_WHITE,name='Arial')
            c.fill=PatternFill('solid',start_color='C00000')
            c.alignment=Alignment(horizontal='center',vertical='center')
            c.border=Border(bottom=Side(style='medium',color='888888'))
        wq.row_dimensions[Rq].height=18; Rq+=1
        for i,d in enumerate(dupes):
            wq.row_dimensions[Rq].height=15
            bg='FFF2F2' if i%2==0 else 'FFFFFF'
            for ci,v in enumerate([d.get('Código',''),d.get('Fecha',''),'2 aplicaciones'],2):
                c=wq.cell(Rq,ci,str(v))
                c.font=Font(size=8,name='Arial')
                c.fill=PatternFill('solid',start_color=bg)
                c.alignment=Alignment(horizontal='left' if ci==2 else 'center',vertical='center',indent=1 if ci==2 else 0)
                c.border=Border(bottom=Side(style='thin',color='DDDDDD'))
            Rq+=1
    else:
        wq.row_dimensions[Rq].height=20
        wq.merge_cells(f'B{Rq}:E{Rq}')
        c=wq[f'B{Rq}']; c.value='  No se encontraron casos con la misma fecha de aplicación.'
        c.font=Font(italic=True,size=9,color='1A6632',name='Arial')
        c.fill=PatternFill('solid',start_color='E2F0D9')
        c.alignment=Alignment(horizontal='left',vertical='center',indent=2)

    # ── Hoja 5: Por Centro ────────────────────────────────────────────────────
    wp=wb.create_sheet('Por Centro'); wp.sheet_properties.tabColor=C_MID
    wp.sheet_view.showGridLines=False
    wp.column_dimensions['A'].width=2; wp.column_dimensions['B'].width=42
    for ltr,w in zip('CDEFGH',[16,16,14,14,14,16]):
        wp.column_dimensions[ltr].width=w
    wp.row_dimensions[1].height=32
    wp.merge_cells('B1:H1')
    ct=wp['B1']; ct.value='🏥  RESUMEN POR CENTRO / SERVICIO DE TRATAMIENTO'
    ct.font=Font(bold=True,size=13,color=C_WHITE,name='Arial')
    ct.fill=PatternFill('solid',start_color=C_DARK)
    ct.alignment=Alignment(horizontal='left',vertical='center',indent=1)
    wp.row_dimensions[2].height=6

    col_centro_wide=None
    for c in wide.columns:
        nc=_norm_str(c)
        if any(k in nc for k in ['codigo del centro','servicio de tratamiento']) and 'trabajo' not in nc:
            col_centro_wide=c; break

    hdrs_pc=['Centro / Servicio de Tratamiento','Aplicaciones','Pacientes únicos',
             'Con IRT2','Sin IRT2\n(pendientes)','Con IRT3','Vals. corregidos']
    cols_pc=list('BCDEFGH')
    wp.row_dimensions[3].height=22
    for col,hdr in zip(cols_pc,hdrs_pc):
        c=wp[f'{col}3']; c.value=hdr
        c.font=Font(bold=True,size=9,color=C_WHITE,name='Arial')
        c.fill=PatternFill('solid',start_color=C_DARK)
        c.alignment=Alignment(horizontal='center' if col!='B' else 'left',
                              vertical='center',wrap_text=True,indent=1 if col=='B' else 0)
        c.border=Border(bottom=Side(style='medium',color='888888'))

    if col_centro_wide:
        resumen_pc=[]
        for centro,grp_w in wide.groupby(col_centro_wide,dropna=False):
            pacs_u   =len(grp_w)
            n_irt2_c =int((grp_w['Tiene_IRT2']=='Sí').sum()) if 'Tiene_IRT2' in grp_w.columns else 0
            n_sin_c  =pacs_u-n_irt2_c
            n_irt3_c =int((grp_w['Tiene_IRT3']=='Sí').sum()) if 'Tiene_IRT3' in grp_w.columns else 0
            n_al_c   =sum(1 for a in alertas if str(a.get('Centro',''))[:60]==str(centro)[:60])
            resumen_pc.append({'Centro':str(centro)[:60] if pd.notna(centro) else '(sin código)',
                               'Aplicaciones':pacs_u,'Pacientes únicos':pacs_u,
                               'Con IRT2':n_irt2_c,'Sin IRT2':n_sin_c,
                               'Con IRT3':n_irt3_c,'Vals':n_al_c})
        df_pc=pd.DataFrame(resumen_pc).sort_values('Aplicaciones',ascending=False)
        for ri,row_p in enumerate(df_pc.itertuples(index=False),4):
            wp.row_dimensions[ri].height=16
            bg='EEF4FB' if ri%2==0 else C_WHITE
            vals=[row_p.Centro,row_p.Aplicaciones,getattr(row_p,'Pacientes únicos',row_p._2),
                  row_p._3,row_p._4,row_p._5,row_p._6]
            for col,v in zip(cols_pc,vals):
                c=wp[f'{col}{ri}']; c.value=v
                es_err=col=='H' and isinstance(v,(int,float)) and int(v)>0
                c.font=Font(size=8,name='Arial',bold=es_err,color='C00000' if es_err else '000000')
                c.fill=PatternFill('solid',start_color=bg)
                c.alignment=Alignment(horizontal='left' if col=='B' else 'center',
                                      vertical='center',indent=1 if col=='B' else 0)
                c.border=Border(bottom=Side(style='thin',color=C_BDR))
        R_tot=4+len(df_pc); wp.row_dimensions[R_tot].height=18
        tots=['TOTAL',df_pc['Aplicaciones'].sum(),df_pc['Pacientes únicos'].sum(),
              df_pc['Con IRT2'].sum(),df_pc['Sin IRT2'].sum(),
              df_pc['Con IRT3'].sum(),df_pc['Vals'].sum()]
        for col,v in zip(cols_pc,tots):
            c=wp[f'{col}{R_tot}']; c.value=v
            c.font=Font(bold=True,size=9,color=C_WHITE,name='Arial')
            c.fill=PatternFill('solid',start_color=C_DARK)
            c.alignment=Alignment(horizontal='left' if col=='B' else 'center',
                                  vertical='center',indent=1 if col=='B' else 0)
    else:
        wp.merge_cells('B3:H3')
        c=wp['B3']; c.value='  No se detectó columna de centro en esta base de datos.'
        c.font=Font(italic=True,size=9,color='595959',name='Arial')
        c.fill=PatternFill('solid',start_color='F2F2F2')
        c.alignment=Alignment(horizontal='left',vertical='center',indent=2)

    # ── Hoja 6: Pendientes IRT2 ───────────────────────────────────────────────
    if 'Alerta_IRT2' in wide.columns:
        _pend=wide[wide['Alerta_IRT2'].isin(['🟠 60-89 dias','🔴 90+ dias'])].copy()
        if len(_pend)>0:
            _cols_p=[]
            if col_centro_wide: _cols_p.append(col_centro_wide)
            _cols_p.append(COL_CODIGO)
            if col_f1: _cols_p.append(col_f1)
            if 'Dias_desde_IRT1' in wide.columns: _cols_p.append('Dias_desde_IRT1')
            _cols_p.append('Alerta_IRT2')
            _tab=_pend[[c for c in _cols_p if c in _pend.columns]].copy()
            _ren={COL_CODIGO:'Código Paciente','Dias_desde_IRT1':'Días desde IRT1','Alerta_IRT2':'Alerta'}
            if col_centro_wide: _ren[col_centro_wide]='Centro / Servicio'
            if col_f1: _ren[col_f1]='Fecha IRT1'
            _tab=_tab.rename(columns={k:v for k,v in _ren.items() if k in _tab.columns})
            sort_cols=['Alerta']+(['Días desde IRT1'] if 'Días desde IRT1' in _tab.columns else [])
            _tab['_ord']=_tab['Alerta'].apply(lambda x:0 if '90' in str(x) else 1)
            _tab=_tab.sort_values(['_ord']+sort_cols[1:],ascending=True).drop(columns='_ord').reset_index(drop=True)

            wp2=wb.create_sheet('Pendientes IRT2')
            wp2.sheet_properties.tabColor='C00000'; wp2.sheet_view.showGridLines=False
            nc=len(_tab.columns)
            wp2.merge_cells(start_row=1,start_column=1,end_row=1,end_column=nc)
            import datetime as _dt
            ct=wp2.cell(1,1)
            ct.value=f'PENDIENTES IRT2  ·  {len(_tab)} pacientes requieren evaluación  ·  {_dt.date.today().strftime("%d/%m/%Y")}'
            ct.font=Font(bold=True,size=12,color=C_WHITE,name='Arial')
            ct.fill=PatternFill('solid',start_color='C00000')
            ct.alignment=Alignment(horizontal='center',vertical='center')
            wp2.row_dimensions[1].height=28
            wp2.merge_cells(start_row=2,start_column=1,end_row=2,end_column=nc)
            cs=wp2.cell(2,1)
            cs.value=f'🔴 {n_rojo} con 90+ días (URGENTE)   |   🟠 {n_naranja} con 60–89 días (PRÓXIMOS)'
            cs.font=Font(size=10,color='444444',name='Arial')
            cs.fill=PatternFill('solid',start_color='FFF2CC')
            cs.alignment=Alignment(horizontal='center',vertical='center')
            wp2.row_dimensions[2].height=20
            for ci,col in enumerate(_tab.columns,1):
                c=wp2.cell(3,ci); c.value=col
                c.font=Font(bold=True,size=9,color=C_WHITE,name='Arial')
                c.fill=PatternFill('solid',start_color='2F3640')
                c.alignment=Alignment(horizontal='center',vertical='center')
                c.border=Border(bottom=Side(style='medium',color='888888'))
                ltr=get_column_letter(ci)
                wp2.column_dimensions[ltr].width=(36 if any(k in col for k in ['Centro','Código']) else
                                                   14 if 'Fecha' in col else 16 if 'Días' in col else
                                                   16 if 'Alerta' in col else 20)
            wp2.row_dimensions[3].height=22
            for ri,row in _tab.iterrows():
                er=ri+4; alv=row.get('Alerta',''); es_r='90' in str(alv)
                bgf=PatternFill('solid',start_color='FDECEA' if es_r else 'FEF3E2')
                for ci,col in enumerate(_tab.columns,1):
                    c=wp2.cell(er,ci); val=row[col]
                    c.value=None if (not isinstance(val,str) and pd.isna(val)) else (int(val) if isinstance(val,float) and val==int(val) else val)
                    c.font=Font(size=9,name='Arial',bold=(col=='Alerta'),
                                color=('C00000' if es_r else 'E67E22') if col=='Alerta' else '222222')
                    c.fill=bgf
                    c.alignment=Alignment(horizontal='center',vertical='center')
                wp2.row_dimensions[er].height=18
            wp2.freeze_panes='A4'

    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf
    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf
